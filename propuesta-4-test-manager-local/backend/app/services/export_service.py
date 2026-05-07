from __future__ import annotations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from ..models import UserStory, TestCase

STATUS_COLORS = {
    "pass": "C6EFCE",
    "fail": "FFC7CE",
    "blocked": "FFEB9C",
    "pending": "DDEBF7",
    "na": "EDEDED",
}

PRIORITY_COLORS = {
    "critical": "FF0000",
    "high": "FF6600",
    "medium": "FFA500",
    "low": "008000",
}

# PatternFill precalculados — evitar crear objetos nuevos en cada iteración del loop
_STATUS_FILLS = {
    status: PatternFill(start_color=color, end_color=color, fill_type="solid")
    for status, color in STATUS_COLORS.items()
}
_STATUS_FILLS_DEFAULT = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _header_cell(ws, row, col, value, fill=None, font_size=11, bold=True, font_color="FFFFFF"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=font_color, size=font_size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    return cell


def _data_cell(ws, row, col, value, fill=None, wrap=True, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    return cell


def export_project_to_excel(project_name: str, stories: list[UserStory]) -> bytes:
    wb = Workbook()

    # ── Cover sheet ───────────────────────────────────────────────────────────
    ws_cover = wb.active
    ws_cover.title = "Resumen"
    ws_cover.column_dimensions["A"].width = 30
    ws_cover.column_dimensions["B"].width = 20

    ws_cover.merge_cells("A1:E1")
    title_cell = ws_cover["A1"]
    title_cell.value = f"Matriz de Casos de Prueba — {project_name}"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_cover.row_dimensions[1].height = 40

    headers = ["Historia de Usuario", "Casos Totales", "Aprobados", "Fallidos", "Pendientes"]
    for i, h in enumerate(headers, 1):
        _header_cell(ws_cover, 2, i, h, fill=SUBHEADER_FILL)
        ws_cover.column_dimensions[get_column_letter(i)].width = 28

    row = 3
    for story in stories:
        cases = story.test_cases
        passed = sum(1 for c in cases if c.status == "pass")
        failed = sum(1 for c in cases if c.status == "fail")
        pending = sum(1 for c in cases if c.status == "pending")
        fill = ALT_ROW_FILL if row % 2 == 0 else None
        _data_cell(ws_cover, row, 1, story.title, fill)
        _data_cell(ws_cover, row, 2, len(cases), fill)
        _data_cell(ws_cover, row, 3, passed, fill)
        _data_cell(ws_cover, row, 4, failed, fill)
        _data_cell(ws_cover, row, 5, pending, fill)
        row += 1

    # ── Test matrix sheet ─────────────────────────────────────────────────────
    ws_matrix = wb.create_sheet("Matriz de Pruebas")

    columns = [
        ("ID", 10), ("Historia de Usuario", 35), ("ID HU Externo", 15),
        ("Título del Caso", 40), ("Tipo", 14), ("Prioridad", 12),
        ("Precondición", 35), ("Pasos", 50), ("Resultado Esperado", 35),
        ("Resultado Actual", 35), ("Estado", 12), ("Notas", 30),
    ]

    for i, (header, width) in enumerate(columns, 1):
        _header_cell(ws_matrix, 1, i, header, fill=HEADER_FILL)
        ws_matrix.column_dimensions[get_column_letter(i)].width = width
    ws_matrix.row_dimensions[1].height = 30

    case_row = 2
    case_num = 1
    for story in stories:
        for tc in story.test_cases:
            steps_text = ""
            if tc.steps:
                for step in tc.steps:
                    if isinstance(step, dict):
                        steps_text += f"{step.get('order', '')}. {step.get('action', '')}\n"
                        if step.get("expected"):
                            steps_text += f"   → {step.get('expected')}\n"

            status_fill = _STATUS_FILLS.get(tc.status, _STATUS_FILLS_DEFAULT)
            row_fill = ALT_ROW_FILL if case_row % 2 == 0 else None

            _data_cell(ws_matrix, case_row, 1, f"TC-{case_num:03d}", row_fill, bold=True)
            _data_cell(ws_matrix, case_row, 2, story.title, row_fill)
            _data_cell(ws_matrix, case_row, 3, story.external_id or "", row_fill)
            _data_cell(ws_matrix, case_row, 4, tc.title, row_fill)
            _data_cell(ws_matrix, case_row, 5, tc.test_type or "functional", row_fill)

            priority_cell = ws_matrix.cell(row=case_row, column=6, value=tc.priority)
            priority_cell.font = Font(
                bold=True,
                color=PRIORITY_COLORS.get(tc.priority, "000000"),
                size=10,
            )
            priority_cell.alignment = Alignment(vertical="top", wrap_text=True)
            priority_cell.border = THIN_BORDER
            if row_fill:
                priority_cell.fill = row_fill

            _data_cell(ws_matrix, case_row, 7, tc.precondition or "", row_fill)
            _data_cell(ws_matrix, case_row, 8, steps_text.strip(), row_fill)
            _data_cell(ws_matrix, case_row, 9, tc.expected_result or "", row_fill)
            _data_cell(ws_matrix, case_row, 10, tc.actual_result or "", row_fill)

            status_cell = ws_matrix.cell(row=case_row, column=11, value=tc.status.upper())
            status_cell.font = Font(bold=True, size=10)
            status_cell.alignment = Alignment(horizontal="center", vertical="top")
            status_cell.border = THIN_BORDER
            status_cell.fill = status_fill

            _data_cell(ws_matrix, case_row, 12, tc.notes or "", row_fill)

            ws_matrix.row_dimensions[case_row].height = max(40, len(steps_text.split("\n")) * 15)
            case_row += 1
            case_num += 1

    ws_matrix.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
